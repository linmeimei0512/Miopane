from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter

from analyze_utils.employee_analyze import EmployeeAnalyze
from utils.write_excel_cell_utils import WriteExcelCellUtils
from utils.dictionary_key import DictionaryKey

class LeaveListAddIDConverter:
    # leave list
    _leave_list_excel_path = None
    _new_leave_list_excel_path = None
    _leave_list_workbook: Workbook = None
    _leave_list_worksheet: Worksheet = None

    # employee list
    _employee_excel_path = None
    _employee_analyze: EmployeeAnalyze = None

    # leave employee list
    _leave_employee_excel_path = None
    _leave_employee_analyze: EmployeeAnalyze = None


    def __init__(self, leave_list_excel_path, employee_excel_path, leave_employee_excel_path, new_leave_list_excel_path=None):
        self._leave_list_excel_path = leave_list_excel_path
        self._employee_excel_path = employee_excel_path
        self._leave_employee_excel_path = leave_employee_excel_path
        self._new_leave_list_excel_path = new_leave_list_excel_path

        self._init_employee_analyze()
        self._init_leave_employee_analyze()


    def _init_employee_analyze(self):
        self._employee_analyze = EmployeeAnalyze(employee_excel_path=self._employee_excel_path)


    def _init_leave_employee_analyze(self):
        self._leave_employee_analyze = EmployeeAnalyze(employee_excel_path=self._leave_employee_excel_path)


    def convert(self):
        self._leave_list_workbook = load_workbook(self._leave_list_excel_path)
        self._leave_list_worksheet = self._leave_list_workbook.worksheets[0]

        row_data_list = list(self._leave_list_worksheet.rows)
        row_data_list.pop(0)    # remove title
        write_list = []
        for index, row_data in enumerate(row_data_list):
            data = [cell.value for cell in row_data]
            name = data[0]
            department_name = data[1]
            id = data[2]
            seniority = data[3]

            # search by leave employee list
            employee_list = self._leave_employee_analyze.search_by_name(name=name)
            # employee_list = []
            if len(employee_list) == 0:
                # search by employee list
                employee_list = self._employee_analyze.search_by_name(name=name)

            # check department
            if department_name is not None:
                employee_list = list(filter(lambda employee: employee['所屬單位'] == department_name, employee_list))

            if len(employee_list) != 0:
                department_name = ', '.join(employee['所屬單位'] for employee in employee_list)
                id = ', '.join(str(employee['工號']) for employee in employee_list)
                seniority = ', '.join(str(employee['內部年資']) for employee in employee_list)

            # print(data, '->', name, department_name, id)

            if len(employee_list) > 1:
                write_list.append({DictionaryKey.VALUE: department_name if department_name is not None else '',
                                   DictionaryKey.START: 'B' + str(index + 2),
                                   DictionaryKey.WIDTH: 30,
                                   DictionaryKey.TEXT_COLOR: 'FF0000'})
                write_list.append({DictionaryKey.VALUE: id if id is not None else '',
                                   DictionaryKey.START: 'C' + str(index + 2),
                                   DictionaryKey.TEXT_COLOR: 'FF0000'})
                write_list.append({DictionaryKey.VALUE: seniority if seniority is not None else '',
                                   DictionaryKey.START: 'D' + str(index + 2),
                                   DictionaryKey.TEXT_COLOR: 'FF0000'})
            else:
                write_list.append({DictionaryKey.VALUE: department_name if department_name is not None else '',
                                   DictionaryKey.START: 'B' + str(index + 2),
                                   DictionaryKey.WIDTH: 30})
                write_list.append({DictionaryKey.VALUE: id if id is not None else '',
                                   DictionaryKey.START: 'C' + str(index + 2)})
                write_list.append({DictionaryKey.VALUE: seniority if seniority is not None else '',
                                   DictionaryKey.START: 'D' + str(index + 2)})

        WriteExcelCellUtils.Write(sheet=self._leave_list_worksheet,
                                  value_list=write_list)
        self._leave_list_workbook.save(filename=self._new_leave_list_excel_path)




if __name__ == '__main__':
    leave_list_add_id_convert = LeaveListAddIDConverter(leave_list_excel_path='../Document/2月實際離職.xlsx',
                                                        employee_excel_path='../Document/目前人力配置.xlsx',
                                                        leave_employee_excel_path='../Document/離職名單_20240313 (2).xlsx',
                                                        new_leave_list_excel_path='../Document/2月實際離職(new).xlsx')
    leave_list_add_id_convert.convert()