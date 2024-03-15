import sys
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter

sys.path.append('..')
from company_sheet import CompanySheet, Company, Department, DepartmentGroup
from analyze_utils.employee_analyze import EmployeeAnalyze
from analyze_utils.real_check_in_and_change_analyze import ReadCheckInAndChangeAnalyze
from analyze_utils.real_leave_analyze import RealLeaveAnalyze
from analyze_utils.expect_check_in_leave_analyze import ExpectCheckInLeaveAnalyze
from utils.dictionary_key import DictionaryKey
from utils.write_excel_cell_utils import WriteExcelCellUtils

class HumanResourceMain(CompanySheet):
    # root cell list
    _root_cell_list = [{DictionaryKey.VALUE: '人力資源報表', DictionaryKey.START: 'A1', DictionaryKey.END: 'C1', DictionaryKey.CENTER: True, DictionaryKey.HEIGHT: 50, DictionaryKey.BOLD: True, DictionaryKey.THIN_BORDER: True},
                       {DictionaryKey.VALUE: '月', DictionaryKey.START: 'D1', DictionaryKey.CENTER: True, DictionaryKey.BOLD: True, DictionaryKey.THIN_BORDER: True},
                       {DictionaryKey.VALUE: '野椿企業集團', DictionaryKey.START: 'A2', DictionaryKey.END: 'D2', DictionaryKey.CENTER: True, DictionaryKey.BOLD: True, DictionaryKey.THIN_BORDER: True},
                       {DictionaryKey.VALUE: '單位異動', DictionaryKey.START: 'K2', DictionaryKey.END: 'L2', DictionaryKey.CENTER: True, DictionaryKey.BOLD: True, DictionaryKey.THIN_BORDER: True}]

    # common cell list
    _common_cell_list = [{DictionaryKey.VALUE: '部門', DictionaryKey.START: 'A3', DictionaryKey.END: 'A4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH:22, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '單位', DictionaryKey.START: 'B3', DictionaryKey.END: 'B4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '預估人力配置', DictionaryKey.START: 'C3', DictionaryKey.END: 'D3', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '正職', DictionaryKey.START: 'C4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '工讀生', DictionaryKey.START: 'D4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '目前人力配置', DictionaryKey.START: 'E3', DictionaryKey.END: 'F3', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '正職', DictionaryKey.START: 'E4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '工讀生', DictionaryKey.START: 'F4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '人力溢 / 缺額', DictionaryKey.START: 'G3', DictionaryKey.END: 'H3', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '正職', DictionaryKey.START: 'G4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '工讀生', DictionaryKey.START: 'H4', DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '月初總人數', DictionaryKey.START: 'I3', DictionaryKey.END: 'I4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '當月入職人數', DictionaryKey.START: 'J3', DictionaryKey.END: 'J4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '轉入單位人數', DictionaryKey.START: 'K3', DictionaryKey.END: 'K4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '轉出單位人數', DictionaryKey.START: 'L3', DictionaryKey.END: 'L4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '當月離職人數', DictionaryKey.START: 'M3', DictionaryKey.END: 'M4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '月末總人數', DictionaryKey.START: 'N3', DictionaryKey.END: 'N4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'DDEBF7'},
                         {DictionaryKey.VALUE: '處理情形', DictionaryKey.START: 'O3', DictionaryKey.END: 'O4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True},
                         {DictionaryKey.VALUE: '預計\n報到/轉調', DictionaryKey.START: 'P3', DictionaryKey.END: 'P4', DictionaryKey.CENTER: True, DictionaryKey.BOLD: True, DictionaryKey.TEXT_COLOR: 'FF0000', DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFFF00'},
                         {DictionaryKey.VALUE: '預計\n離職/轉調', DictionaryKey.START: 'Q3', DictionaryKey.END: 'Q4', DictionaryKey.CENTER: True, DictionaryKey.BOLD: True, DictionaryKey.TEXT_COLOR: 'FF0000', DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFFF00'},
                         {DictionaryKey.VALUE: '備註', DictionaryKey.START: 'R3', DictionaryKey.END: 'R4', DictionaryKey.CENTER: True, DictionaryKey.WIDTH: 14, DictionaryKey.THIN_BORDER: True}]

    _common_prompt_cell_list = [{DictionaryKey.VALUE: '預估人力配置：依照各店平均業績所預估之人力', DictionaryKey.START: 'A23', DictionaryKey.END: 'S23', DictionaryKey.THIN_BORDER: True},
                                {DictionaryKey.VALUE: '目前人力配置：當月月底之在職人數=月初總人數＋當月入職人數 - 當月離職人數', DictionaryKey.START: 'A24', DictionaryKey.END: 'S24', DictionaryKey.THIN_BORDER: True},
                                {DictionaryKey.VALUE: '月初總人數：當月1號在職人數，『不包含』1號到職之團員', DictionaryKey.START: 'A25', DictionaryKey.END: 'S25', DictionaryKey.THIN_BORDER: True},
                                {DictionaryKey.VALUE: '當月入職人數：全月到職之團員人數', DictionaryKey.START: 'A26', DictionaryKey.END: 'S26', DictionaryKey.THIN_BORDER: True},
                                {DictionaryKey.VALUE: '當月離職人數：全月離職之團員人數', DictionaryKey.START: 'A27', DictionaryKey.END: 'S27', DictionaryKey.THIN_BORDER: True},
                                {DictionaryKey.VALUE: '當月離職率：當月離職人數 / (月初總人數＋當月入職人數)＊％', DictionaryKey.START: 'A28', DictionaryKey.END: 'S28', DictionaryKey.THIN_BORDER: True}]

    # company
    _company = Company(name='', sheet_name='集團總表')
    _company.department_list = [Department(name='中央辦公室', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=SUM(\'總公司\'!C4:C12)',
                                                                                                     number_current_full_time_employee_operator='=SUM(\'總公司\'!E4:E12)',
                                                                                                     new_check_in_operator='=SUM(\'總公司\'!J4:J12)',
                                                                                                     change_in_operator='=SUM(\'總公司\'!K4:K12)',
                                                                                                     change_out_operator='=SUM(\'總公司\'!L4:L12)',
                                                                                                     real_leave_employee_operator='=SUM(\'總公司\'!M4:M12)',
                                                                                                     expect_check_in_operator='=SUM(\'總公司\'!P4:P12)',
                                                                                                     expect_leave_operator='=SUM(\'總公司\'!Q4:Q12)')]),
                                Department(name='中央醬料', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'總公司\'!C13',
                                                                                                   number_current_full_time_employee_operator='=\'總公司\'!E13',
                                                                                                   number_expect_part_time_employee_operator='=\'總公司\'!D13',
                                                                                                   number_current_part_time_employee_operator='=\'總公司\'!F13',
                                                                                                   new_check_in_operator='=\'總公司\'!J13',
                                                                                                   change_in_operator='=\'總公司\'!K13',
                                                                                                   change_out_operator='=\'總公司\'!L13',
                                                                                                   real_leave_employee_operator='=\'總公司\'!M13',
                                                                                                   expect_check_in_operator='=\'總公司\'!P13',
                                                                                                   expect_leave_operator='=\'總公司\'!Q13')]),
                                Department(name='中央麵包', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'總公司\'!C14',
                                                                                                   number_current_full_time_employee_operator='=\'總公司\'!E14',
                                                                                                   number_expect_part_time_employee_operator='=\'總公司\'!D14',
                                                                                                   number_current_part_time_employee_operator='=\'總公司\'!F14',
                                                                                                   new_check_in_operator='=\'總公司\'!J14',
                                                                                                   change_in_operator='=\'總公司\'!K14',
                                                                                                   change_out_operator='=\'總公司\'!L14',
                                                                                                   real_leave_employee_operator='=\'總公司\'!M14',
                                                                                                   expect_check_in_operator='=\'總公司\'!P14',
                                                                                                   expect_leave_operator='=\'總公司\'!Q14')]),
                                Department(name='中央西點', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'總公司\'!C15',
                                                                                                   number_current_full_time_employee_operator='=\'總公司\'!E15',
                                                                                                   number_expect_part_time_employee_operator='=\'總公司\'!D15',
                                                                                                   number_current_part_time_employee_operator='=\'總公司\'!F15',
                                                                                                   new_check_in_operator='=\'總公司\'!J15',
                                                                                                   change_in_operator='=\'總公司\'!K15',
                                                                                                   change_out_operator='=\'總公司\'!L15',
                                                                                                   real_leave_employee_operator='=\'總公司\'!M15',
                                                                                                   expect_check_in_operator='=\'總公司\'!P15',
                                                                                                   expect_leave_operator='=\'總公司\'!Q15')]),
                                Department(name='中央電商', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'總公司\'!C16',
                                                                                                   number_current_full_time_employee_operator='=\'總公司\'!E16',
                                                                                                   number_expect_part_time_employee_operator='=\'總公司\'!D16',
                                                                                                   number_current_part_time_employee_operator='=\'總公司\'!F16',
                                                                                                   new_check_in_operator='=\'總公司\'!J16',
                                                                                                   change_in_operator='=\'總公司\'!K16',
                                                                                                   change_out_operator='=\'總公司\'!L16',
                                                                                                   real_leave_employee_operator='=\'總公司\'!M16',
                                                                                                   expect_check_in_operator='=\'總公司\'!P16',
                                                                                                   expect_leave_operator='=\'總公司\'!Q16')]),
                                Department(name='廚房物流', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'總公司\'!C17',
                                                                                                   number_current_full_time_employee_operator='=\'總公司\'!E17',
                                                                                                   number_expect_part_time_employee_operator='=\'總公司\'!D17',
                                                                                                   number_current_part_time_employee_operator='=\'總公司\'!F17',
                                                                                                   new_check_in_operator='=\'總公司\'!J17',
                                                                                                   change_in_operator='=\'總公司\'!K17',
                                                                                                   change_out_operator='=\'總公司\'!L17',
                                                                                                   real_leave_employee_operator='=\'總公司\'!M17',
                                                                                                   expect_check_in_operator='=\'總公司\'!P17',
                                                                                                   expect_leave_operator='=\'總公司\'!Q17')]),
                                Department(name='宅配物流', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'總公司\'!C18',
                                                                                                   number_current_full_time_employee_operator='=\'總公司\'!E18',
                                                                                                   number_expect_part_time_employee_operator='=\'總公司\'!D18',
                                                                                                   number_current_part_time_employee_operator='=\'總公司\'!F18',
                                                                                                   new_check_in_operator='=\'總公司\'!J18',
                                                                                                   change_in_operator='=\'總公司\'!K18',
                                                                                                   change_out_operator='=\'總公司\'!L18',
                                                                                                   real_leave_employee_operator='=\'總公司\'!M18',
                                                                                                   expect_check_in_operator='=\'總公司\'!P18',
                                                                                                   expect_leave_operator='=\'總公司\'!Q18')]),
                                Department(name=''),
                                Department(name='Miacucina', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'Miacucina\'!D20',
                                                                                                    number_current_full_time_employee_operator='=\'Miacucina\'!F20',
                                                                                                    number_expect_part_time_employee_operator='=\'Miacucina\'!E20',
                                                                                                    number_current_part_time_employee_operator='=\'Miacucina\'!G20',
                                                                                                    new_check_in_operator='=\'Miacucina\'!K20',
                                                                                                    change_in_operator='=\'Miacucina\'!L20',
                                                                                                    change_out_operator='=\'Miacucina\'!M20',
                                                                                                    real_leave_employee_operator='=\'Miacucina\'!N20',
                                                                                                    expect_check_in_operator='=\'Miacucina\'!Q20',
                                                                                                    expect_leave_operator='=\'Miacucina\'!R20')]),
                                Department(name='Miopane', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'Miopane\'!D20',
                                                                                                  number_current_full_time_employee_operator='=\'Miopane\'!F20',
                                                                                                  number_expect_part_time_employee_operator='=\'Miopane\'!E20',
                                                                                                  number_current_part_time_employee_operator='=\'Miopane\'!G20',
                                                                                                  new_check_in_operator='=\'Miopane\'!K20',
                                                                                                  change_in_operator='=\'Miopane\'!L20',
                                                                                                  change_out_operator='=\'Miopane\'!M20',
                                                                                                  real_leave_employee_operator='=\'Miopane\'!N20',
                                                                                                  expect_check_in_operator='=\'Miopane\'!Q20',
                                                                                                  expect_leave_operator='=\'Miopane\'!R20')]),
                                Department(name='餐廳其他品牌', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'餐廳其他品牌\'!D9',
                                                                                                       number_current_full_time_employee_operator='=\'餐廳其他品牌\'!F9',
                                                                                                       number_expect_part_time_employee_operator='=\'餐廳其他品牌\'!E9',
                                                                                                       number_current_part_time_employee_operator='=\'餐廳其他品牌\'!G9',
                                                                                                       new_check_in_operator='=\'餐廳其他品牌\'!K9',
                                                                                                       change_in_operator='=\'餐廳其他品牌\'!L9',
                                                                                                       change_out_operator='=\'餐廳其他品牌\'!M9',
                                                                                                       real_leave_employee_operator='=\'餐廳其他品牌\'!N9',
                                                                                                       expect_check_in_operator='=\'餐廳其他品牌\'!Q9',
                                                                                                       expect_leave_operator='=\'餐廳其他品牌\'!R9')]),
                                Department(name='Dreamers', department_group_list=[DepartmentGroup(name='', number_expect_full_time_employee_operator='=\'Dreamers\'!D19',
                                                                                                   number_current_full_time_employee_operator='=\'Dreamers\'!F19',
                                                                                                   number_expect_part_time_employee_operator='=\'Dreamers\'!E19',
                                                                                                   number_current_part_time_employee_operator='=\'Dreamers\'!G19',
                                                                                                   new_check_in_operator='=\'Dreamers\'!K19',
                                                                                                   change_in_operator='=\'Dreamers\'!L19',
                                                                                                   change_out_operator='=\'Dreamers\'!M19',
                                                                                                   real_leave_employee_operator='=\'Dreamers\'!N19',
                                                                                                   expect_check_in_operator='=\'Dreamers\'!Q19',
                                                                                                   expect_leave_operator='=\'Dreamers\'!R19')])
                                ]




    def __init__(self, human_resource_workbook: Workbook, month,
                 employee_analyze: EmployeeAnalyze,
                 real_check_in_and_change_analyze: ReadCheckInAndChangeAnalyze,
                 real_leave_analyze: RealLeaveAnalyze,
                 expect_check_in_leave_analyze: ExpectCheckInLeaveAnalyze):
        self._root_cell_list[1][DictionaryKey.VALUE] = str(month) + '月'
        self._department_value_start_column = 'C'
        self._department_start_row = 5
        super().__init__(human_resource_workbook=human_resource_workbook,
                         month=month,
                         employee_analyze=employee_analyze,
                         real_check_in_and_change_analyze=real_check_in_and_change_analyze,
                         real_leave_analyze=real_leave_analyze,
                         expect_check_in_leave_analyze=expect_check_in_leave_analyze)

        self._write_department_list()
        self._write_total()


    def _write_department_list(self):
        '''
        Write department list
        '''
        super()._write_department_list()



    def _write_total(self):
        '''
        Write total
        '''
        total_cell_list = []
        total_cell_list.append({DictionaryKey.VALUE: '集團店鋪業績總額', DictionaryKey.START: 'A' + str(self._department_finish_row),
                                DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFF2CC'})
        total_cell_list.append({DictionaryKey.VALUE: '', DictionaryKey.START: 'B' + str(self._department_finish_row),
                                DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFF2CC'})

        for column in range(column_index_from_string('C'), column_index_from_string('O')):
            total_cell_list.append({DictionaryKey.VALUE: '=SUM({}5:{}16)'.format(get_column_letter(column), get_column_letter(column)),
                                    DictionaryKey.START: get_column_letter(column) + str(self._department_finish_row),
                                    DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFF2CC',
                                    DictionaryKey.RULE_TEXT_OPERATOR: 'lessThan', DictionaryKey.RULE_TEXT_FORMULA: '0', DictionaryKey.RULE_TEXT_COLOR: 'FF0000'})

        for column in range(column_index_from_string('P'), column_index_from_string('R')):
            total_cell_list.append({DictionaryKey.VALUE: '=SUM({}5:{}16)'.format(get_column_letter(column), get_column_letter(column)),
                                    DictionaryKey.START: get_column_letter(column) + str(self._department_finish_row),
                                    DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFF2CC',
                                    DictionaryKey.RULE_TEXT_OPERATOR: 'lessThan', DictionaryKey.RULE_TEXT_FORMULA: '0', DictionaryKey.RULE_TEXT_COLOR: 'FF0000'})

        WriteExcelCellUtils.Write(sheet=self._company_sheet,
                                  value_list=total_cell_list)