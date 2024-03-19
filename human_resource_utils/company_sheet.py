from typing import List
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter
from decimal import Decimal, ROUND_HALF_UP

from human_resource_utils.analyze_utils import EmployeeAnalyze, ReadCheckInAndChangeAnalyze, RealLeaveAnalyze, ExpectCheckInLeaveAnalyze
from utils import DictionaryKey, WriteExcelCellUtils


class DepartmentGroup:
    # name
    name = 'department group name'
    search_name = 'department group search name'

    # full-time employee
    number_expect_full_time_employee = 0
    number_current_full_time_employee = 0

    # full-time employee operator
    number_expect_full_time_employee_operator = None
    number_current_full_time_employee_operator = None

    # part-time employee
    number_expect_part_time_employee = -1
    number_current_part_time_employee = -1

    # part-time employee operator
    number_expect_part_time_employee_operator = None
    number_current_part_time_employee_operator = None

    # employee begin of the month
    number_employee_at_begin_of_the_month_operator = None

    # new check in
    new_check_in_operator = None

    # change in and out operator
    change_in_operator = None
    change_out_operator = None

    # real leave employee operator
    real_leave_employee_operator = None

    # expect check in and leave
    expect_check_in_operator = None
    expect_leave_operator = None

    def __init__(self, name, search_name=None,
                 number_expect_full_time_employee=-1, number_expect_full_time_employee_operator=None, number_current_full_time_employee_operator=None,
                 number_expect_part_time_employee=-1, number_expect_part_time_employee_operator=None, number_current_part_time_employee_operator=None,
                 number_employee_at_begin_of_the_month_operator=None,
                 new_check_in_operator=None,
                 change_in_operator=None, change_out_operator=None,
                 real_leave_employee_operator=None,
                 expect_check_in_operator=None, expect_leave_operator=None):
        self.name = name
        self.search_name = search_name
        self.number_expect_full_time_employee = number_expect_full_time_employee
        self.number_expect_full_time_employee_operator = number_expect_full_time_employee_operator
        self.number_current_full_time_employee_operator = number_current_full_time_employee_operator
        self.number_expect_part_time_employee = number_expect_part_time_employee
        self.number_expect_part_time_employee_operator = number_expect_part_time_employee_operator
        self.number_current_part_time_employee_operator = number_current_part_time_employee_operator
        self.number_employee_at_begin_of_the_month_operator = number_employee_at_begin_of_the_month_operator
        self.new_check_in_operator = new_check_in_operator
        self.change_in_operator = change_in_operator
        self.change_out_operator = change_out_operator
        self.real_leave_employee_operator = real_leave_employee_operator
        self.expect_check_in_operator = expect_check_in_operator
        self.expect_leave_operator = expect_leave_operator

    def __str__(self):
        return 'DepartmentGroup( name: {}, search name: {}'.format(self.name, self.search_name) + \
               'number of expect full time employee: {}, number of current full time employee: {}, number of overflow shortage full time employee: {}, '.format(self.number_expect_full_time_employee, self.number_current_full_time_employee, self.number_overflow_shortage_full_time_employee) + \
               'number of expect part time employee: {}, number of current full time employee: {}, number of overflow shortage part time employee: {}'.format(self.number_expect_part_time_employee, self.number_current_part_time_employee, self.number_overflow_shortage_part_time_employee)


class Department:
    # name
    name = 'department name'

    # department list
    department_group_list: List[DepartmentGroup] = []

    def __init__(self, name, department_group_list: List[DepartmentGroup] = []):
        self.name = name
        self.department_group_list = department_group_list

    def __str__(self):
        return 'Department( name: {}, '.format(self.name) + \
               'group list: {}'.format(", ".join(str(group) for group in self.department_group_list))


class Company:
    # name
    name = 'company name'
    sheet_name = 'company sheet name'

    # department list
    department_list: List[Department] = []

    def __init__(self, name, sheet_name, department_list: List[Department] = []):
        self.name = name
        self.sheet_name = sheet_name
        self.department_list = department_list



class CompanySheet(object):
    # human resource workbook
    _human_resource_workbook: Workbook = None

    # month
    _month = None

    # sheet
    _company_sheet: Worksheet = None
    _company_sheet_name = ''

    # root cell list
    _root_cell_list = []

    # common cell list
    _common_cell_list = []
    _common_prompt_cell_list = []

    # department list
    DEPARTMENT_LIST_DEPARTMENT_KEY = 'department'
    DEPARTMENT_LIST_GROUPS_KEY = 'groups'
    _department_start_row = 0
    _department_finish_row = 0
    _department_value_start_column = 'C'
    _department_value_start_column_index = column_index_from_string(_department_value_start_column)
    _department_list = []

    # company
    _company: Company = None

    # employee analyze
    _employee_analyze: EmployeeAnalyze = None

    # real check-in and change analyze
    _real_check_in_and_change_analyze: ReadCheckInAndChangeAnalyze = None

    # real leave analyze
    _real_leave_analyze: RealLeaveAnalyze = None

    # expect check in leave analyze
    _expect_check_in_leave_analyze: ExpectCheckInLeaveAnalyze = None

    # note
    _note = ''
    _note_text_color = None

    # debug
    _debug = False


    def __init__(self, human_resource_workbook: Workbook, month,
                 employee_analyze: EmployeeAnalyze,
                 real_check_in_and_change_analyze: ReadCheckInAndChangeAnalyze,
                 real_leave_analyze: RealLeaveAnalyze,
                 expect_check_in_leave_analyze: ExpectCheckInLeaveAnalyze,
                 debug=False):
        self._human_resource_workbook = human_resource_workbook
        self._month = month
        self._department_value_start_column_index = column_index_from_string(self._department_value_start_column)
        self._employee_analyze = employee_analyze
        self._real_check_in_and_change_analyze = real_check_in_and_change_analyze
        self._real_leave_analyze = real_leave_analyze
        self._expect_check_in_leave_analyze = expect_check_in_leave_analyze
        self._debug = debug

        self._create_company_sheet()
        self._write_root_cell()
        self._write_common_cell()
        self._write_common_prompt()


    def _create_company_sheet(self):
        '''
        Create company sheet
        '''
        self._company_sheet = self._human_resource_workbook.create_sheet(title=self._company.sheet_name)


    def _write_root_cell(self):
        '''
        Write root cell
        '''
        WriteExcelCellUtils.Write(sheet=self._company_sheet,
                                  value_list=self._root_cell_list)


    def _write_common_cell(self):
        '''
        Write common cell
        '''
        WriteExcelCellUtils.Write(sheet=self._company_sheet,
                                  value_list=self._common_cell_list)


    def _write_common_prompt(self):
        '''
        Write common prompt
        '''
        WriteExcelCellUtils.Write(sheet=self._company_sheet,
                                  value_list=self._common_prompt_cell_list)


    def _write_department_list(self):
        '''
        Write department list
        '''
        _department_list = []
        index = 0
        for department in self._company.department_list:
            if department.name == '':
                index += 1
                continue

            _department_title_start = 'A' + str(self._department_start_row + index)

            # group
            for group in department.department_group_list:
                if self._debug:
                    print('\n***** {} - {} *****'.format(department.name, group.name))
                self._note = ''
                self._note_text_color = None
                row_index = self._department_start_row + index
                _department_list.append({DictionaryKey.VALUE: group.name,
                                         DictionaryKey.START: 'B' + str(row_index),
                                         DictionaryKey.WIDTH: 13, DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})


                # full-time employee
                _department_list += self._get_full_time_employee_data(group=group, index=index)

                # part-time employee
                _department_list += self._get_part_time_employee_data(group=group, index=index)

                # new check in
                _department_list += self._get_new_check_in_data(group=group, index=index)

                # real change in and out
                _department_list += self._get_real_change_in_out_data(group=group, index=index)

                # real leave employee
                _department_list += self._get_real_leave(group=group, row_index=row_index)

                # total employee at the end of the month
                _department_list.append({DictionaryKey.VALUE: '=' + self._get_column_string_by_shift_index(shift_index=6) + str(row_index) +
                                                              '+' + self._get_column_string_by_shift_index(shift_index=7) + str(row_index) +
                                                              '+' + self._get_column_string_by_shift_index(shift_index=8) + str(row_index) +
                                                              '-' + self._get_column_string_by_shift_index(shift_index=9) + str(row_index) +
                                                              '-' + self._get_column_string_by_shift_index(shift_index=10) + str(row_index),
                                         DictionaryKey.START: self._get_column_string_by_shift_index(shift_index=11) + str(row_index),
                                         DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

                # expect check in and leave
                _department_list += self._get_expect_check_in_leave(group=group, row_index=row_index)

                # write note
                _department_list.append({DictionaryKey.VALUE: self._note,
                                         DictionaryKey.START: self._get_column_string_by_shift_index(shift_index=15) + str(row_index),
                                         DictionaryKey.THIN_BORDER: True, DictionaryKey.WIDTH: 50, DictionaryKey.TEXT_WRAP: True, DictionaryKey.TEXT_COLOR: self._note_text_color})

                # employee shortage by manager
                _department_list += self._get_employee_shortage_by_manager(group=group, row_index=row_index)

                index += 1

            # department
            _department_list.append({DictionaryKey.VALUE: department.name,
                                     DictionaryKey.START: _department_title_start,
                                     DictionaryKey.END: 'A' + str(self._department_start_row + index - 1),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

            WriteExcelCellUtils.Write(sheet=self._company_sheet,
                                      value_list=_department_list)
            self._department_finish_row = self._department_start_row + index


    def _get_full_time_employee_data(self, group: DepartmentGroup, index):
        '''
        Get full-time employee data

        Args:
            group:
            index:

        Returns:

        '''
        _department_list = []

        # expect
        _expect_full_time_location = self._department_value_start_column + str(self._department_start_row + index)
        _department_list.append({DictionaryKey.VALUE: group.number_expect_full_time_employee if group.number_expect_full_time_employee_operator is None else group.number_expect_full_time_employee_operator,
                                 DictionaryKey.START: _expect_full_time_location,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
        # current
        _current_full_time_employee_list = self._get_employee(department_group_search_name=group.search_name, is_full_time=True)
        group.number_current_full_time_employee = len(_current_full_time_employee_list)
        _comment = '\n'.join(employee['顯示名稱'] for employee in _current_full_time_employee_list)
        _current_full_time_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 2) + str(self._department_start_row + index)
        _department_list.append({DictionaryKey.VALUE: group.number_current_full_time_employee if group.number_current_full_time_employee_operator is None else group.number_current_full_time_employee_operator,
                                 DictionaryKey.COMMENT: _comment if _comment != '' else None,
                                 DictionaryKey.START: _current_full_time_location,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
        # overflow / shortage
        _overflow_shortage_full_time_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 4) + str(self._department_start_row + index)
        _department_list.append({DictionaryKey.VALUE: '={}-{}'.format(_current_full_time_location, _expect_full_time_location),
                                 DictionaryKey.START: _overflow_shortage_full_time_location,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True,
                                 DictionaryKey.RULE_TEXT_OPERATOR: 'lessThan', DictionaryKey.RULE_TEXT_FORMULA: '0', DictionaryKey.RULE_TEXT_COLOR: 'FF0000'})
        return _department_list


    def _get_part_time_employee_data(self, group: DepartmentGroup, index):
        '''
        Get part-time employee data

        Args:
            group:
            index:

        Returns:

        '''
        _department_list = []

        _expect_part_time_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 1) + str(self._department_start_row + index)
        _current_part_time_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 3) + str(self._department_start_row + index)
        _overflow_shortage_part_time_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 5) + str(self._department_start_row + index)

        if group.search_name is not None:
            if group.number_expect_part_time_employee >= 0:
                # expect
                _department_list.append({DictionaryKey.VALUE: group.number_expect_part_time_employee,
                                         DictionaryKey.START: _expect_part_time_location,
                                         DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

                # current
                _current_part_time_employee_list = self._get_employee(department_group_search_name=group.search_name, is_full_time=False)
                _comment = '\n'.join(employee['顯示名稱'] for employee in _current_part_time_employee_list)
                group.number_current_part_time_employee = len(_current_part_time_employee_list)
                _department_list.append({DictionaryKey.VALUE: group.number_current_part_time_employee,
                                         DictionaryKey.COMMENT: _comment if _comment != '' else None,
                                         DictionaryKey.START: _current_part_time_location,
                                         DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

                # overflow / shortage
                _department_list.append({DictionaryKey.VALUE: '={}-{}'.format(_current_part_time_location, _expect_part_time_location),
                                         DictionaryKey.START: _overflow_shortage_part_time_location,
                                         DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True,
                                         DictionaryKey.RULE_TEXT_OPERATOR: 'lessThan', DictionaryKey.RULE_TEXT_FORMULA: '0', DictionaryKey.RULE_TEXT_COLOR: 'FF0000'})
        else:
            # expect
            _department_list.append({DictionaryKey.VALUE: group.number_expect_part_time_employee_operator if group.number_expect_part_time_employee_operator is not None else '',
                                     DictionaryKey.START: _expect_part_time_location,
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

            # current
            _department_list.append({DictionaryKey.VALUE: group.number_current_part_time_employee_operator if group.number_current_part_time_employee_operator is not None else '',
                                     DictionaryKey.START: _current_part_time_location,
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

            # overflow / shortage
            _department_list.append({DictionaryKey.VALUE: '={}-{}'.format(_current_part_time_location, _expect_part_time_location),
                                     DictionaryKey.START: _overflow_shortage_part_time_location,
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True,
                                     DictionaryKey.RULE_TEXT_OPERATOR: 'lessThan', DictionaryKey.RULE_TEXT_FORMULA: '0', DictionaryKey.RULE_TEXT_COLOR: 'FF0000'})
        return _department_list


    def _get_new_check_in_data(self, group: DepartmentGroup, index):
        _department_list = []

        _location = get_column_letter(column_index_from_string(self._department_value_start_column) + 7) + str(self._department_start_row + index)
        _value = ''
        _comment = ''
        if group.search_name is not None:
            new_check_in_list = self._real_check_in_and_change_analyze.search_new_check_in_list(search_group_name=group.search_name)
            _value = len(new_check_in_list)
            _comment = '\n'.join(('FT' if employee['身分類別(後)'] == '正式' else 'PT') + employee['顯示名稱'] for employee in new_check_in_list)

        else:
            _value = group.new_check_in_operator if group.new_check_in_operator is not None else ''

        _department_list.append({DictionaryKey.VALUE: _value,
                                 DictionaryKey.START: _location,
                                 DictionaryKey.COMMENT: _comment if _comment != '' else None,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
        return _department_list



    def _get_real_change_in_out_data(self, group: DepartmentGroup, index):
        '''
        Get change in and out data

        Args:
            group:
            index:

        Returns:

        '''
        _department_list = []

        _in_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 8) + str(self._department_start_row + index)
        _out_location = get_column_letter(column_index_from_string(self._department_value_start_column) + 9) + str(self._department_start_row + index)
        _in_value = ''
        _in_comment = ''
        _out_value = ''
        _out_comment = ''

        if group.search_name is not None:
            # change in
            _change_in_employee_list = self._real_check_in_and_change_analyze.search_change_list(search_group_name=group.search_name, in_or_out='in')
            _in_value = len(_change_in_employee_list)
            _in_comment = '\n'.join('{}{}(原:{}{}))'
                                    .format('FT' if employee['身分類別(後)'] == '正式' else 'PT',
                                            employee['顯示名稱'],
                                            employee['所屬單位(前)'],
                                            'FT' if employee['身分類別(前)'] == '正式' else 'PT')
                                    for employee in _change_in_employee_list)

            # change out
            _change_out_employee_list = self._real_check_in_and_change_analyze.search_change_list(search_group_name=group.search_name, in_or_out='out')
            _out_value = len(_change_out_employee_list)
            _out_comment = '\n'.join(('FT' if employee['身分類別(後)'] == '正式' else 'PT') + employee['顯示名稱'] for employee in _change_out_employee_list)
            _out_comment = '\n'.join('{}{}(後:{}{})'
                                     .format('FT' if employee['身分類別(前)'] == '正式' else 'PT',
                                             employee['顯示名稱'],
                                             employee['所屬單位(後)'],
                                             'FT' if employee['身分類別(後)'] == '正式' else 'PT')
                                     for employee in _change_out_employee_list)

        else:
            # change in
            _in_value = group.change_in_operator

            # change out
            _out_value = group.change_out_operator

        _department_list.append({DictionaryKey.VALUE: _in_value,
                                 DictionaryKey.COMMENT: _in_comment if _in_comment is not '' else None,
                                 DictionaryKey.COMMENT_WIDTH: 300,
                                 DictionaryKey.START: _in_location,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
        _department_list.append({DictionaryKey.VALUE: _out_value,
                                 DictionaryKey.COMMENT: _out_comment if _out_comment is not '' else None,
                                 DictionaryKey.COMMENT_WIDTH: 300,
                                 DictionaryKey.START: _out_location,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
        return _department_list


    def _get_real_leave(self, group: DepartmentGroup, row_index):
        '''
        Get real leave employee

        Args:
            group:
            row_index:

        Returns:

        '''
        _department_list = []
        _location = get_column_letter(self._department_value_start_column_index + 10) + str(row_index)
        _leave_employee_list = []
        _leave_without_pay_employee_list = []

        if group.search_name is not None:
            _leave_employee_list = self._real_leave_analyze.search_leave_list(search_group_name=group.search_name)
            _leave_without_pay_employee_list = self._real_check_in_and_change_analyze.search_leave_without_pay_list(search_group_name=group.search_name)
            print('leave without pay employee list: [{}] -> {}'.format(len(_leave_without_pay_employee_list), _leave_without_pay_employee_list))
            _value = len(_leave_employee_list + _leave_without_pay_employee_list)

        else:
            _value = group.real_leave_employee_operator

        # number of real leave employee
        _department_list.append({DictionaryKey.VALUE: _value,
                                 DictionaryKey.START: _location,
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

        _leave_employee_name = ''
        _leave_employee_seniority = ''

        # leave employee
        _leave_employee_name_list = []
        _leave_employee_seniority_list = []
        if len(_leave_employee_list) > 0:
            # print(_leave_employee_list)
            _leave_employee_name_list = [leave_employee['姓名'] for leave_employee in _leave_employee_list]
            _leave_employee_seniority_list = [str(Decimal(str(leave_employee['內部年資'])).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)) for leave_employee in _leave_employee_list]
            # _leave_employee_seniority_list = [str(np.round(float(leave_employee['內部年資']), 1)) for leave_employee in _leave_employee_list]
            # _leave_employee_name = '\n'.join(leave_employee['姓名'] for leave_employee in _leave_employee_list)
            # _leave_employee_seniority = '\n'.join(str(np.round(float(leave_employee['內部年資']), 1)) for leave_employee in _leave_employee_list)

        # leave without pay employee
        _leave_without_pay_name_list = []
        _leave_without_pay_seniority_list = []
        if len(_leave_without_pay_employee_list) > 0:
            _leave_without_pay_name_list = [leave_without_pay_employee['顯示名稱'] + '(留停)' for leave_without_pay_employee in _leave_without_pay_employee_list]
            _leave_without_pay_seniority_list = [''] * len(_leave_without_pay_employee_list)


        _department_list.append({DictionaryKey.VALUE: '\n'.join(_leave_employee_name_list + _leave_without_pay_name_list),
                                 DictionaryKey.START: get_column_letter(self._department_value_start_column_index + 16) + str(row_index),
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.WIDTH: 12})
        _department_list.append({DictionaryKey.VALUE: '\n'.join(_leave_employee_seniority_list + _leave_without_pay_seniority_list),
                                 DictionaryKey.START: get_column_letter(self._department_value_start_column_index + 17) + str(row_index),
                                 DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

        return _department_list


    def _get_expect_check_in_leave(self, group: DepartmentGroup, row_index):
        _department_list = []
        _expect_check_in_number = 0
        _expect_leave_number = 0

        if group.search_name is not None:
            _expect_leave_list = self._expect_check_in_leave_analyze.search_by_group_name(search_group_name=group.search_name)
            if len(_expect_leave_list) == 0:
                self._note += ('\n' if self._note != '' else '') + '未回覆預計表單'
            else:
                if len(_expect_leave_list) > 1:
                    self._note_text_color = 'FF0000'
                    self._note += ('\n' if self._note != '' else '') + '重複填寫預計表單'

                # check in
                _expect_check_in_full_time_number = _expect_leave_list[0]['預計報到人數-正職團員']
                _expect_check_in_part_time_number = _expect_leave_list[0]['預計報到人數-部分工時團員(PT)']
                _expect_check_in_number = _expect_check_in_full_time_number + _expect_check_in_part_time_number

                # leave
                _expect_leave_full_time_number = _expect_leave_list[0]['預計離職人數-正職團員']
                _expect_leave_part_time_number = _expect_leave_list[0]['預計離職人數-部分工時團員(PT)']
                _expect_leave_number = _expect_leave_full_time_number + _expect_leave_part_time_number

                # write note
                if _expect_check_in_number > 0 or _expect_leave_number > 0:
                    self._note += ('\n' if self._note != '' else '') + '預計'
                    if _expect_check_in_number > 0:
                        self._note += '報到' + \
                                        ('FT*{}'.format(str(_expect_check_in_full_time_number)) if _expect_check_in_full_time_number > 0 else '') + \
                                        ('、' if _expect_check_in_full_time_number > 0 and _expect_check_in_part_time_number > 0 else '') + \
                                        ('PT*{}'.format(str(_expect_check_in_part_time_number)) if _expect_check_in_part_time_number > 0 else '')
                    if _expect_leave_number > 0:
                        self._note += (', ' if _expect_check_in_number > 0 else '') + \
                                        '離職' + \
                                        ('FT*{}'.format(str(_expect_leave_full_time_number)) if _expect_leave_full_time_number > 0 else '') + \
                                        ('、' if _expect_leave_full_time_number > 0 and _expect_leave_part_time_number > 0 else '') + \
                                        ('PT*{}'.format(str(_expect_leave_part_time_number)) if _expect_leave_part_time_number > 0 else '')

            _department_list.append({DictionaryKey.VALUE: _expect_check_in_number,
                                     DictionaryKey.START: self._get_column_string_by_shift_index(13) + str(row_index),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
            _department_list.append({DictionaryKey.VALUE: _expect_leave_number,
                                     DictionaryKey.START: self._get_column_string_by_shift_index(14) + str(row_index),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

        else:
            _department_list.append({DictionaryKey.VALUE: group.expect_check_in_operator if group.expect_check_in_operator is not None else '',
                                     DictionaryKey.START: self._get_column_string_by_shift_index(13) + str(row_index),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
            _department_list.append({DictionaryKey.VALUE: group.expect_leave_operator if group.expect_leave_operator is not None else '',
                                     DictionaryKey.START: self._get_column_string_by_shift_index(14) + str(row_index),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

        return _department_list


    def _get_employee_shortage_by_manager(self, group: DepartmentGroup, row_index):
        _department_list = []
        _full_time_shortage = 0
        _part_time_shortage = 0

        if group.search_name is not None:
            _employee_shortage_list = self._expect_check_in_leave_analyze.search_by_group_name(search_group_name=group.search_name)
            if len(_employee_shortage_list) > 0:
                _full_time_shortage = _employee_shortage_list[0]['人力缺額-正職人員']
                _part_time_shortage = _employee_shortage_list[0]['人力缺額-部分工時人員(PT)']

            _department_list.append({DictionaryKey.VALUE: _full_time_shortage,
                                     DictionaryKey.START: self._get_column_string_by_shift_index(19) + str(row_index),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})
            _department_list.append({DictionaryKey.VALUE: _part_time_shortage,
                                     DictionaryKey.START: self._get_column_string_by_shift_index(20) + str(row_index),
                                     DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True})

        return _department_list


    def _get_employee(self, department_group_search_name, is_full_time):
        '''
        Get employee

        Args:
            department_group_search_name:
            is_full_time:

        Returns:

        '''
        _employee_list = self._employee_analyze.search_by_department_group(search_group_name=department_group_search_name,
                                                                           is_full_time=is_full_time)
        _leave_employee_list = self._real_leave_analyze.search_leave_list(search_group_name=department_group_search_name)
        _leave_without_pay_employee_list = self._real_check_in_and_change_analyze.search_leave_without_pay_list(search_group_name=department_group_search_name)

        employee_list = []
        for employee in _employee_list:
            _leave = False

            # leave
            for leave_employee in _leave_employee_list:
                if employee['工號'] == leave_employee['工號']:
                    _leave = True
                    self._add_note(note='{}({},{}) 已離職，但還在總表中'.format(employee['顯示名稱'], employee['工號'], employee['所屬單位']))
                    # self._note += ('\n' if self._note != '' else '') + '{}({},{}) 已離職，但還在總表中'.format(employee['顯示名稱'], employee['工號'], employee['所屬單位'])

            # leave without pay
            for leave_without_pay_employee in _leave_without_pay_employee_list:
                if employee['工號'] == leave_without_pay_employee['工號(前)']:
                    _leave = True
                    self._add_note(note='{}({},{}) 留職停薪'.format(employee['顯示名稱'], employee['工號'], employee['所屬單位']))

            if not _leave:
                employee_list.append(employee)

        return employee_list


    def _get_column_string_by_shift_index(self, shift_index):
        return get_column_letter(self._department_value_start_column_index + shift_index)


    def _write_total(self, start_column, end_column):
        '''
        Write total

        Args:
            start_column:
            end_column:

        Returns:

        '''
        _total_list = [{DictionaryKey.VALUE: '合計',
                        DictionaryKey.START: 'B' + str(self._department_finish_row),
                        DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True}]

        for column in range(column_index_from_string(start_column), column_index_from_string(end_column) + 1):
            _total_list.append({DictionaryKey.VALUE: '=SUM({}:{})'.format(get_column_letter(column) + str(self._department_start_row), get_column_letter(column) + str(self._department_finish_row - 1)),
                                DictionaryKey.START: get_column_letter(column) + str(self._department_finish_row),
                                DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFF2CC',
                                DictionaryKey.RULE_TEXT_OPERATOR: 'lessThan', DictionaryKey.RULE_TEXT_FORMULA: '0', DictionaryKey.RULE_TEXT_COLOR: 'FF0000'})

        for column in range(column_index_from_string(end_column) + 2, column_index_from_string(end_column) + 4):
            _total_list.append({DictionaryKey.VALUE: '=SUM({}:{})'.format(get_column_letter(column) + str(self._department_start_row), get_column_letter(column) + str(self._department_finish_row - 1)),
                                DictionaryKey.START: get_column_letter(column) + str(self._department_finish_row),
                                DictionaryKey.CENTER: True, DictionaryKey.THIN_BORDER: True, DictionaryKey.BACKGROUND_COLOR: 'FFF2CC'})

        WriteExcelCellUtils.Write(sheet=self._company_sheet,
                                  value_list=_total_list)


    def _add_note(self, note):
        self._note += ('\n' if self._note != '' else '') + note