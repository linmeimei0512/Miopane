import sys
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formatting.rule import CellIsRule

sys.path.append('../utils')
from dictionary_key import DictionaryKey

class WriteExcelCellUtils:

    @staticmethod
    def Write(sheet: Worksheet, value_list: list):
        for value in value_list:
            _start = coordinate_to_tuple(value[DictionaryKey.START])
            _start_row = _start[0]
            _start_column = _start[1]
            _end = coordinate_to_tuple(value[DictionaryKey.END]) if DictionaryKey.END in value.keys() else None
            _end_row = _end[0] if _end is not None else _start_row
            _end_column = _end[1] if _end is not None else _start_column
            _value = value[DictionaryKey.VALUE] if DictionaryKey.VALUE in value.keys() else None

            _center = value[DictionaryKey.CENTER] if DictionaryKey.CENTER in value.keys() else False
            _width = value[DictionaryKey.WIDTH] if DictionaryKey.WIDTH in value.keys() else None
            _height = value[DictionaryKey.HEIGHT] if DictionaryKey.HEIGHT in value.keys() else None

            _bold = value[DictionaryKey.BOLD] if DictionaryKey.BOLD in value.keys() else False
            _text_color = value[DictionaryKey.TEXT_COLOR] if DictionaryKey.TEXT_COLOR in value.keys() else '000000'
            _text_wrap = value[DictionaryKey.TEXT_WRAP] if DictionaryKey.TEXT_WRAP in value.keys() else None

            _thin_border = value[DictionaryKey.THIN_BORDER] if DictionaryKey.THIN_BORDER in value.keys() else False

            _background_color = value[DictionaryKey.BACKGROUND_COLOR] if DictionaryKey.BACKGROUND_COLOR in value.keys() else None

            _rule_operator = value[DictionaryKey.RULE_TEXT_OPERATOR] if DictionaryKey.RULE_TEXT_OPERATOR in value.keys() else None
            _rule_formula = value[DictionaryKey.RULE_TEXT_FORMULA] if DictionaryKey.RULE_TEXT_FORMULA in value.keys() else None
            _rule_text_color = value[DictionaryKey.RULE_TEXT_COLOR] if DictionaryKey.RULE_TEXT_COLOR in value.keys() else None

            # print('value: ', _value)

            # write cell value
            cell = sheet.cell(row=_start_row, column=_start_column, value=_value)

            # merge row and column
            if _start_row != _end_row and _start_column != _end_column:
                sheet.merge_cells(start_row=_start_row, start_column=_start_column, end_row=_end_row, end_column=_end_column)

            elif _start_row != _end_row:
                sheet.merge_cells(start_row=_start_row, start_column=_start_column, end_row=_end_row, end_column=_start_column)

            elif _start_column != _end_column:
                sheet.merge_cells(start_row=_start_row, start_column=_start_column, end_row=_start_row, end_column=_end_column)

            # center
            if _center:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif _text_wrap is not None:
                cell.alignment = Alignment(wrap_text=_text_wrap)

            # width
            if _width is not None:
                sheet.column_dimensions[get_column_letter(_start_column)].width = _width

            # height
            if _height is not None:
                sheet.row_dimensions[_start_row].height = _height

            # text bold, color
            cell.font = Font(bold=_bold, color=_text_color)

            # thin border
            if _thin_border:
                for row in range(_start_row, _end_row + 1):
                    for column in range(_start_column, _end_column + 1):
                        sheet.cell(row=row, column=column).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # background color
            if _background_color is not None:
                for row in range(_start_row, _end_row + 1):
                    for column in range(_start_column, _end_column + 1):
                        sheet.cell(row=row, column=column).fill = PatternFill(start_color=_background_color, end_color=_background_color, fill_type='solid')

            # rule
            if _rule_operator is not None:
                for row in range(_start_row, _end_row + 1):
                    for column in range(_start_column, _end_column + 1):
                        sheet.conditional_formatting.add('{}:{}'.format(get_column_letter(column) + str(row), get_column_letter(column) + str(row)),
                                                         CellIsRule(operator=_rule_operator,
                                                                    formula=[_rule_formula],
                                                                    stopIfTrue=True,
                                                                    font=Font(color=_rule_text_color)))
